# PDF 加载器
from langchain_community.document_loaders import PyPDFLoader
from rich import print
from pathlib import Path
from pypdf import PdfWriter
from langchain_core.documents import Document

"""
loader = PyPDFLoader("handbook.pdf")
# 一页一个Document，所以len(docs)=PDF页数
docs = loader.load()
print("Document数量", len(docs))
for doc in docs:
    print(doc)

## 新建一个PDF写入器
# writer = PdfWriter()
## 加两个空白页 尺寸A4
# writer.add_blank_page(width=595, height=842)
# writer.add_blank_page(width=595, height=842)
# with open("scan.pdf", "wb") as f:
#    writer.write(f)

loader = PyPDFLoader("scan.pdf")
docs = loader.load()
for doc in docs:
    print(doc)

from langchain_community.document_loaders import Docx2txtLoader

loader = Docx2txtLoader("contract.docx")
docs = loader.load()
for doc in docs:
    print(doc)


from langchain_community.document_loaders import CSVLoader

# Path("orders.csv").write_text(
#    "order_id,status,eta\nA1000,已发货,明天\nA1001,运输中,后天\n", encoding="utf-8"
# )

loader = CSVLoader("orders.csv", encoding="utf-8")
docs = loader.load()
for doc in docs:
    print(doc)


from openpyxl import Workbook
import openpyxl


# 创建工作簿
# wb = Workbook()
## 获取默认的工作表
# ws = wb.active
# ws.title = "订单"  # type: ignore
# ws.append(["order_id", "status", "time"])  # type: ignore
# ws.append(["A1000", "已发货", "明天"])  # type: ignore
# ws.append(["A1001", "运输中", "后天"])  # type: ignore
# wb.create_sheet()
# wb.save("orders.xlsx")
def load_excel(path: Path) -> list[Document]:
    ""把Excel每一行读成一个Document,首行当表头""
    # path  要读取的excel文件路径
    # read_only=True 让大文件也能低内存读取
    # data_only=True 表示读公式计算的结果，而不是读取公式本身
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    # 收集所有的产出的Document
    out: list[Document] = []
    # 一个Excel里面可能包含多个工作表，所以需要遍历
    for ws in wb.worksheets:
        # values_only=True表示只读单元格的值，不要样式对象
        rows = ws.iter_rows(values_only=True)
        # 读取第一行当表头，next第2个参数是取不值的默认值
        header = next(rows, None)
        if not header:
            continue
        # 遍历剩下的数据行，i从0开始的
        for i, row in enumerate(rows):
            # 如果整行所有的列都是None,跳过这一行
            if all(cell is None for cell in row):
                continue
            # 拼成 列名:值 的格式，和CSVLoader一致
            body = "\n".join(f"{h}: {v}" for h, v in zip(header, row) if h is not None)
            out.append(
                Document(
                    page_content=body,
                    metadata={"source": path.as_posix(), "sheet": ws.title, "row": i},
                )
            )
    wb.close()
    return out


docs = load_excel(Path("orders.xlsx"))
for doc in docs:
    print(doc)
"""
import os

os.environ.setdefault("USER_AGENT", "Chrome/16.0")
from langchain_community.document_loaders import WebBaseLoader

loader = WebBaseLoader(["https://example.com"])
docs = loader.load()
for doc in docs:
    print(doc)
